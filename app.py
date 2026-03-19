import streamlit as st
import xml.etree.ElementTree as ET
from datetime import datetime
import pandas as pd
import requests
import warnings
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from zeep import Client

warnings.filterwarnings("ignore")

# ==========================================
# CONSTANTES
# ==========================================
TASA_CONTRAPRESTACION = 0.05
MONEDAS_MXN = {"MXN"}
MONEDAS_USD = {"USD"}

# ==========================================
# TIPO DE CAMBIO (Caché para no saturar la API)
# ==========================================
@st.cache_data(ttl=3600)
def obtener_tipo_cambio_apis() -> float:
    try:
        url = "https://www.banxico.org.mx/SieAPIRest/service/v1/series/SF43718/datos/oportuno"
        r = requests.get(url, headers={"Bmx-Token": ""}, timeout=5)
        return float(r.json()["bmx"]["series"]["datos"]["dato"].replace(",", ""))
    except Exception:
        pass
    try:
        r = requests.get("https://open.er-api.com/v6/latest/USD", timeout=5)
        return float(r.json()["rates"]["MXN"])
    except Exception:
        return 18.0000 # Valor de respaldo si no hay internet

# ==========================================
# EXTRACCIÓN CFDI
# ==========================================
def extraer_datos_cfdi(archivo_xml) -> dict:
    try:
        archivo_xml.seek(0)
        root = ET.parse(archivo_xml).getroot()

        def get(*attrs):
            return next((root.get(a) for a in attrs if root.get(a) is not None), None)

        sello = get('Sello', 'sello') or ""
        total = get('Total', 'total')

        rfc_emisor = nombre_emisor = None
        rfc_receptor = nombre_receptor = None
        uuid = None
        descripciones = []

        for elem in root.iter():
            tag = elem.tag.lower()
            if tag.endswith('emisor') and not rfc_emisor:
                rfc_emisor    = elem.get('Rfc') or elem.get('rfc')
                nombre_emisor = elem.get('Nombre') or elem.get('nombre')
            elif tag.endswith('receptor') and not rfc_receptor:
                rfc_receptor    = elem.get('Rfc') or elem.get('rfc')
                nombre_receptor = elem.get('Nombre') or elem.get('nombre')
            elif tag.endswith('timbrefiscaldigital') and not uuid:
                uuid = elem.get('UUID') or elem.get('uuid')
            elif tag.endswith('concepto'):
                desc = elem.get('Descripcion') or elem.get('descripcion')
                if desc:
                    descripciones.append(desc.replace('\n', ' ').replace('\r', '').strip())

        link_sat = ""
        if all([rfc_emisor, rfc_receptor, total, uuid]):
            link_sat = (
                f"https://verificacfdi.facturaelectronica.sat.gob.mx/default.aspx"
                f"?id={uuid}&re={rfc_emisor}&rr={rfc_receptor}&tt={total}&fe={sello[-8:]}"
            )

        return {
            'Archivo':             archivo_xml.name,
            'Serie':               get('Serie', 'serie'),
            'Folio':               get('Folio', 'folio'),
            'UUID':                uuid,
            'RFC_Emisor':          rfc_emisor,
            'Nombre_Emisor':       nombre_emisor,
            'RFC_Receptor':        rfc_receptor,
            'Nombre_Receptor':     nombre_receptor,
            'Moneda':              get('Moneda', 'moneda'),
            'Descripcion':         " | ".join(descripciones) or "Sin descripción",
            'SubTotal':            get('SubTotal', 'subTotal'),
            'Total':               total,
            'Fecha_Emision':       get('Fecha', 'fecha'),
            'Link_Validacion_SAT': link_sat,
            'Error_Lectura':       None,
        }
    except Exception as e:
        return {'Archivo': archivo_xml.name, 'Error_Lectura': str(e)}

# ==========================================
# CONSULTA SAT
# ==========================================
def consultar_sat(rfc_emisor, rfc_receptor, total, uuid) -> tuple[str, str]:
    wsdl = 'https://consultaqr.facturaelectronica.sat.gob.mx/ConsultaCFDIService.svc?wsdl'
    try:
        client = Client(wsdl)
        resp = client.service.Consulta(f"?re={rfc_emisor}&rr={rfc_receptor}&tt={total}&id={uuid}")
        return resp['Estado'], resp['CodigoEstatus']
    except Exception as e:
        return 'Error de conexión', str(e)

# ==========================================
# INTERFAZ WEB (STREAMLIT)
# ==========================================
st.set_page_config(page_title="Auditor de Facturas Portuarias", layout="wide")

st.title("🚢 Auditor de Facturas y Contraprestación")
st.markdown("Valida los XML contra el SAT, revisa periodos y calcula el equivalente en MXN de forma automática.")

# Controles de Configuración
col1, col2, col3 = st.columns()
with col1:
    fecha_inicio = st.date_input("Fecha de Inicio del periodo")
with col2:
    fecha_fin = st.date_input("Fecha de Fin del periodo")
with col3:
    tc_auto = obtener_tipo_cambio_apis()
    tipo_cambio = st.number_input("Tipo de cambio USD→MXN aplicable", value=tc_auto, format="%.4f")

archivos_subidos = st.file_uploader("Arrastra aquí los archivos XML", type=['xml'], accept_multiple_files=True)

if st.button("🚀 Iniciar Auditoría", type="primary") and archivos_subidos:
    resultados = []
    barra_progreso = st.progress(0)
    texto_estado = st.empty()
    total_archivos = len(archivos_subidos)

    for i, archivo in enumerate(archivos_subidos):
        texto_estado.text(f"Procesando: {archivo.name} ({i+1}/{total_archivos})")
        datos = extraer_datos_cfdi(archivo)

        if not datos.get('Error_Lectura'):
            # Validar rango
            try:
                fecha_dt = datetime.strptime(datos['Fecha_Emision'], '%Y-%m-%dT%H:%M:%S').date()
                dentro = fecha_inicio <= fecha_dt <= fecha_fin
                datos['Periodo_Correcto'] = 'Sí' if dentro else 'No (Fuera de rango)'
                datos['Periodo_Servicio'] = fecha_dt.strftime("%Y-%m") if dentro else ""
            except (ValueError, TypeError):
                datos['Periodo_Correcto'] = 'Error en formato de fecha'
                datos['Periodo_Servicio'] = ""

            # Consulta SAT
            if all([datos.get(k) for k in ('RFC_Emisor', 'RFC_Receptor', 'Total', 'UUID')]):
                estado, codigo = consultar_sat(datos['RFC_Emisor'], datos['RFC_Receptor'], datos['Total'], datos['UUID'])
            else:
                estado, codigo = 'Datos incompletos', 'N/A'
            datos['Estado_SAT'] = estado
            datos['Codigo_SAT'] = codigo

            # Montos y divisas
            moneda = (datos.get('Moneda') or '').upper().strip()
            subtotal = pd.to_numeric(datos.get('SubTotal'), errors='coerce') or 0.0
            total_v = pd.to_numeric(datos.get('Total'), errors='coerce') or 0.0

            if moneda in MONEDAS_MXN:
                datos['SubTotal_MXN'] = subtotal
                datos['Total_MXN'] = total_v
                datos['SubTotal_USD'] = None
                datos['Total_USD'] = None
                datos['Tipo_Cambio_Aplicado'] = None
                datos['SubTotal_MXN_Equivalente'] = subtotal
            elif moneda in MONEDAS_USD:
                datos['SubTotal_MXN'] = None
                datos['Total_MXN'] = None
                datos['SubTotal_USD'] = subtotal
                datos['Total_USD'] = total_v
                datos['Tipo_Cambio_Aplicado'] = tipo_cambio
                datos['SubTotal_MXN_Equivalente'] = subtotal * tipo_cambio
            else:
                datos['SubTotal_MXN'] = subtotal
                datos['Total_MXN'] = total_v
                datos['SubTotal_USD'] = None
                datos['Total_USD'] = None
                datos['Tipo_Cambio_Aplicado'] = None
                datos['SubTotal_MXN_Equivalente'] = subtotal
                datos['Moneda'] = f"{moneda} ⚠"

            # Contraprestación
            datos['Contraprestacion_5%_MXN'] = datos['SubTotal_MXN_Equivalente'] * TASA_CONTRAPRESTACION

        resultados.append(datos)
        barra_progreso.progress((i + 1) / total_archivos)

    texto_estado.text("¡Auditoría completada! Aplicando formato al reporte...")

    # ==========================================
    # PREPARAR DATAFRAME
    # ==========================================
    df = pd.DataFrame(resultados)
    
    columnas_ordenadas = [
        'Archivo', 'Serie', 'Folio', 'UUID', 'RFC_Emisor', 'Nombre_Emisor',
        'RFC_Receptor', 'Nombre_Receptor', 'Moneda', 'SubTotal_MXN', 'Total_MXN',
        'SubTotal_USD', 'Total_USD', 'Tipo_Cambio_Aplicado', 'SubTotal_MXN_Equivalente',
        'Contraprestacion_5%_MXN', 'Descripcion', 'Fecha_Emision', 'Periodo_Correcto',
        'Periodo_Servicio', 'Estado_SAT', 'Codigo_SAT', 'Link_Validacion_SAT', 'Error_Lectura'
    ]
    df = df[[col for col in columnas_ordenadas if col in df.columns]]

    # ==========================================
    # CREAR EXCEL EN MEMORIA CON FORMATO
    # ==========================================
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Detalle")
        wb = writer.book
        ws = writer.sheets["Detalle"]

        # Estilos
        GRIS_HEADER = PatternFill("solid", fgColor="2E4057")
        VERDE_CLARO = PatternFill("solid", fgColor="E2EFDA")
        AMARILLO = PatternFill("solid", fgColor="FFF2CC")
        THIN = Side(style='thin', color="CCCCCC")
        BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        headers = {cell.value: cell.column for cell in ws}
        col_periodo_ok = headers.get('Periodo_Correcto')
        col_periodo_sv = headers.get('Periodo_Servicio')

        # Formato de Encabezados
        for cell in ws:
            cell.fill = GRIS_HEADER
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDE
        ws.row_dimensions.height = 35

        # Formato de Celdas y Condicionales
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            dentro = col_periodo_ok and row[col_periodo_ok - 1].value == 'Sí'
            fuera = col_periodo_ok and row[col_periodo_ok - 1].value == 'No (Fuera de rango)'

            for cell in row:
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="center")
                cell.border = BORDE
                if dentro:
                    cell.fill = VERDE_CLARO
                elif fuera:
                    cell.fill = AMARILLO

            if col_periodo_sv:
                c = row[col_periodo_sv - 1]
                if not c.value:
                    c.fill = PatternFill("solid", fgColor="FFD966")
                    c.font = Font(name="Arial", size=9, italic=True, color="7F6000")

            for col_nombre in ('SubTotal_MXN', 'Total_MXN', 'SubTotal_USD', 'Total_USD',
                               'SubTotal_MXN_Equivalente', 'Contraprestacion_5%_MXN'):
                col_idx = headers.get(col_nombre)
                if col_idx and row[col_idx - 1].value is not None:
                    row[col_idx - 1].number_format = '#,##0.00'

        # Fila de Totales
        max_row = ws.max_row
        fila_total = max_row + 2
        ws.cell(fila_total, 1, "TOTALES").font = Font(bold=True, name="Arial", size=10)

        for col_nombre in ('SubTotal_MXN', 'Total_MXN', 'SubTotal_USD', 'Total_USD',
                           'SubTotal_MXN_Equivalente', 'Contraprestacion_5%_MXN'):
            col_idx = headers.get(col_nombre)
            if col_idx:
                letra = get_column_letter(col_idx)
                c = ws.cell(fila_total, col_idx)
                c.value = f"=SUM({letra}2:{letra}{max_row})"
                c.font = Font(bold=True, name="Arial", size=10)
                c.number_format = '#,##0.00'
                c.border = BORDE

        # Ajuste de Columnas
        anchos = {
            'Archivo': 25, 'Serie': 8, 'Folio': 8, 'UUID': 38,
            'RFC_Emisor': 16, 'Nombre_Emisor': 30, 'Moneda': 10, 'Descripcion': 45,
            'SubTotal_MXN': 16, 'Total_MXN': 16, 'SubTotal_USD': 16, 'Total_USD': 16,
            'SubTotal_MXN_Equivalente': 22, 'Contraprestacion_5%_MXN': 22,
            'Fecha_Emision': 18, 'Periodo_Correcto': 16, 'Periodo_Servicio': 16,
            'Estado_SAT': 16, 'Codigo_SAT': 14, 'Link_Validacion_SAT': 20
        }
        for nombre, ancho in anchos.items():
            if nombre in headers:
                ws.column_dimensions[get_column_letter(headers[nombre])].width = ancho

        ws.freeze_panes = "A2"

    datos_excel = output.getvalue()

    # ==========================================
    # MÉTRICAS Y DESCARGA WEB
    # ==========================================
    st.divider()
    validas = df[df['Estado_SAT'].str.startswith('Vigente', na=False)]
    
    st.subheader(f"📊 Resumen del Periodo | TC Aplicado: ${tipo_cambio:,.4f}")
    c1, c2, c3 = st.columns(3)
    c1.metric("Facturas Procesadas", len(df))
    c2.metric("Total Equivalente (MXN)", f"${df['SubTotal_MXN_Equivalente'].sum():,.2f}")
    c3.metric("Bolsa Contraprestación (5%)", f"${df['Contraprestacion_5%_MXN'].sum():,.2f}")

    st.download_button(
        label="📥 Descargar Reporte Formateado (Excel)",
        data=datos_excel,
        file_name=f"Auditoria_{fecha_inicio}_al_{fecha_fin}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
